"""FastAPI backend for Enigma Normalizer.

In-process file cache: dicts keyed by uploaded filename.
No auth, no logging, CORS open for http://localhost:5173.
"""
from __future__ import annotations

import csv as csv_module
import io
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

from normalizers import LABELS, get_normalizer
from utils.anomalies import scan_anomalies
from utils.detect import scan_dataframe
from utils.text_extract import extract_document, rebuild_document
from utils.text_scan import apply_replacements, group_by_type, scan_text_document

app = FastAPI(title="Enigma Normalizer API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# In-memory caches
# ---------------------------------------------------------------------------

# filename -> {sheets_data, is_csv, normalized_bytes, mapping_payload}
_file_cache: dict[str, dict[str, Any]] = {}

# filename -> {doc: ExtractedDocument, matches: list | None}
_doc_cache: dict[str, dict[str, Any]] = {}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _read_excel(data: bytes) -> dict[str, pd.DataFrame]:
    for engine in ("calamine", "openpyxl"):
        try:
            xls = pd.ExcelFile(io.BytesIO(data), engine=engine)
            return {name: pd.read_excel(xls, sheet_name=name, dtype=object) for name in xls.sheet_names}
        except Exception:
            continue
    xls = pd.ExcelFile(io.BytesIO(data))
    return {name: pd.read_excel(xls, sheet_name=name, dtype=object) for name in xls.sheet_names}


def _read_csv(data: bytes) -> dict[str, pd.DataFrame]:
    sample = data[:4096].decode("utf-8-sig", errors="replace")
    sep = ";" if sample.count(";") >= sample.count(",") else ","
    encodings = ("utf-8-sig", "utf-8", "cp1251", "latin-1")

    for encoding in encodings:
        try:
            df = pd.read_csv(io.BytesIO(data), sep=sep, dtype=object, encoding=encoding, on_bad_lines="skip")
            return {"Sheet1": df}
        except Exception:
            continue

    for encoding in encodings:
        try:
            df = pd.read_csv(
                io.BytesIO(data), sep=sep, dtype=object, encoding=encoding,
                engine="python", quoting=csv_module.QUOTE_NONE, on_bad_lines="skip",
            )
            return {"Sheet1": df}
        except Exception:
            continue

    for encoding in encodings:
        try:
            df = pd.read_csv(
                io.BytesIO(data), sep=sep, dtype=object, encoding=encoding,
                engine="python", quoting=csv_module.QUOTE_NONE, escapechar="\\", on_bad_lines="skip",
            )
            return {"Sheet1": df}
        except Exception:
            continue

    raise ValueError("Не удалось прочитать CSV-файл.")


def _build_mapping_excel(payload: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")

    for sh_name, sh_data in payload["sheets"].items():
        ws = wb.create_sheet(title=sh_name[:31])
        headers = ["Колонка", "Тип данных", "Каноническое значение", "Варианты", "Кол-во вхождений", "Уверенность"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row_idx = 2
        for col_name, col_data in sh_data.get("per_column", {}).items():
            for group in col_data.get("groups", []):
                ws.cell(row=row_idx, column=1, value=col_name)
                ws.cell(row=row_idx, column=2, value=col_data.get("data_type_label", ""))
                ws.cell(row=row_idx, column=3, value=group["canonical"])
                ws.cell(row=row_idx, column=4, value=" | ".join(group["variants"]))
                ws.cell(row=row_idx, column=5, value=group["count"])
                ws.cell(row=row_idx, column=6, value=group["confidence"])
                row_idx += 1

        for col_cells in ws.columns:
            length = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(length + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _require_file(filename: str) -> dict:
    entry = _file_cache.get(filename)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"Файл '{filename}' не найден в кэше. Загрузите заново.")
    return entry


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class ScanRequest(BaseModel):
    filename: str
    sheet: str


class ColumnSpec(BaseModel):
    name: str
    data_type: str


class RunRequest(BaseModel):
    filename: str
    sheet: str
    columns: list[ColumnSpec]


class GroupSpec(BaseModel):
    canonical: str
    variants: list[str]
    apply: bool


class SheetNormalizeSpec(BaseModel):
    columns: list[ColumnSpec]
    groups: dict[str, list[GroupSpec]]


class NormalizeRequest(BaseModel):
    filename: str
    sheets: dict[str, SheetNormalizeSpec]


class AnomaliesRequest(BaseModel):
    filename: str
    sheets: list[str]
    sample_size: int | None = None


class DocScanRequest(BaseModel):
    filename: str


class SelectionSpec(BaseModel):
    canonical: str
    variants: list[str]
    apply: bool


class DocNormalizeRequest(BaseModel):
    filename: str
    selections: dict[str, list[SelectionSpec]]


# ---------------------------------------------------------------------------
# Excel/CSV endpoints
# ---------------------------------------------------------------------------

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in (".xlsx", ".xls", ".csv"):
        raise HTTPException(status_code=400, detail="Поддерживаются только .xlsx, .xls, .csv")

    data = await file.read()
    is_csv = ext == ".csv"

    try:
        sheets_data = _read_csv(data) if is_csv else _read_excel(data)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

    filename = file.filename
    _file_cache[filename] = {
        "sheets_data": sheets_data,
        "is_csv": is_csv,
        "normalized_bytes": None,
        "mapping_payload": None,
    }

    return {
        "filename": filename,
        "is_csv": is_csv,
        "sheets": {
            name: {
                "columns": [str(c) for c in df.columns],
                "row_count": len(df),
            }
            for name, df in sheets_data.items()
        },
    }


@app.post("/api/scan")
async def scan_sheet(req: ScanRequest):
    entry = _require_file(req.filename)
    df: pd.DataFrame | None = entry["sheets_data"].get(req.sheet)
    if df is None:
        raise HTTPException(status_code=404, detail=f"Лист '{req.sheet}' не найден")

    scans = scan_dataframe(df)
    return {
        "scans": [
            {
                "column": s.column,
                "detected_type": s.detected_type,
                "confidence": s.confidence,
                "recommended": s.recommended,
                "scores": s.scores,
            }
            for s in scans
        ]
    }


@app.post("/api/run")
async def run_normalizers(req: RunRequest):
    entry = _require_file(req.filename)
    df: pd.DataFrame | None = entry["sheets_data"].get(req.sheet)
    if df is None:
        raise HTTPException(status_code=404, detail=f"Лист '{req.sheet}' не найден")

    results: dict[str, list[dict]] = {}
    for col_spec in req.columns:
        col, data_type = col_spec.name, col_spec.data_type
        if col not in df.columns:
            raise HTTPException(status_code=422, detail=f"Колонка '{col}' не найдена в листе")
        try:
            normalizer = get_normalizer(data_type)
        except ValueError as e:
            raise HTTPException(status_code=422, detail=str(e))

        values = [str(v) for v in df[col].dropna().tolist()]
        candidates = normalizer.build_candidates(values)
        results[col] = [
            {
                "canonical": c.canonical,
                "variants": c.variants,
                "count": c.count,
                "confidence": round(c.confidence, 3),
                "meta": c.meta,
            }
            for c in candidates
        ]

    return {"results": results}


@app.post("/api/normalize")
async def normalize(req: NormalizeRequest):
    entry = _require_file(req.filename)
    sheets_data: dict[str, pd.DataFrame] = entry["sheets_data"]

    grand_total_changed = 0
    sheets_payload: dict[str, dict] = {}
    normalized_by_sheet: dict[str, pd.DataFrame] = {}

    for sh_name, sh_spec in req.sheets.items():
        df = sheets_data.get(sh_name)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Лист '{sh_name}' не найден")

        normalized_df = df.copy()
        per_column_payload: dict[str, dict] = {}
        sheet_changed = 0

        col_type_map = {c.name: c.data_type for c in sh_spec.columns}

        for col_name, group_specs in sh_spec.groups.items():
            data_type = col_type_map.get(col_name)
            mapping: dict[str, str] = {}
            applied_groups = []

            for g in group_specs:
                if not g.apply:
                    continue
                canonical = (g.canonical or "").strip()
                if not canonical:
                    continue
                for v in g.variants:
                    mapping[v] = canonical
                applied_groups.append({
                    "canonical": canonical,
                    "variants": g.variants,
                    "count": len(g.variants),
                    "confidence": 1.0,
                })

            if col_name not in normalized_df.columns:
                continue

            col_series = normalized_df[col_name]
            str_series = col_series.astype(str)
            if mapping:
                replaced = str_series.map(mapping)
                new_series = replaced.where(replaced.notna(), col_series)
            else:
                new_series = col_series

            changed = int((str_series != new_series.astype(str)).sum())
            sheet_changed += changed
            normalized_df[col_name] = new_series

            per_column_payload[col_name] = {
                "data_type": data_type,
                "data_type_label": LABELS.get(data_type, data_type) if data_type else "",
                "values_changed": changed,
                "mapping": mapping,
                "groups": applied_groups,
            }

        normalized_by_sheet[sh_name] = normalized_df
        grand_total_changed += sheet_changed
        sheets_payload[sh_name] = {
            "columns": list(sh_spec.groups.keys()),
            "values_changed": sheet_changed,
            "per_column": per_column_payload,
        }

    mapping_payload = {
        "meta": {
            "source_file": req.filename,
            "sheets": list(sheets_payload.keys()),
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "total_values_changed": grand_total_changed,
        },
        "sheets": sheets_payload,
    }

    # Build and cache normalized xlsx
    xlsx_buf = io.BytesIO()
    with pd.ExcelWriter(xlsx_buf, engine="xlsxwriter") as writer:
        for sh_name, df in sheets_data.items():
            (normalized_by_sheet.get(sh_name, df)).to_excel(writer, sheet_name=sh_name, index=False)

    entry["normalized_bytes"] = xlsx_buf.getvalue()
    entry["normalized_by_sheet"] = normalized_by_sheet
    entry["mapping_payload"] = mapping_payload

    return {
        "mapping_payload": mapping_payload,
        "changed_total": grand_total_changed,
    }


@app.get("/api/download/normalized")
async def download_normalized(filename: str = Query(...), format: str = Query("xlsx")):
    entry = _require_file(filename)
    normalized_bytes: bytes | None = entry.get("normalized_bytes")
    if normalized_bytes is None:
        raise HTTPException(status_code=400, detail="Сначала выполните нормализацию через POST /api/normalize")

    base = Path(filename).stem
    if format == "csv":
        norm_by_sheet: dict[str, pd.DataFrame] = entry.get("normalized_by_sheet", {})
        if not norm_by_sheet:
            raise HTTPException(status_code=400, detail="Нет нормализованных данных")
        df_csv = next(iter(norm_by_sheet.values()))
        csv_bytes = df_csv.to_csv(index=False, sep=";").encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{base}__normalized.csv"'},
        )

    return StreamingResponse(
        io.BytesIO(normalized_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{base}__normalized.xlsx"'},
    )


@app.get("/api/download/mapping")
async def download_mapping(filename: str = Query(...)):
    entry = _require_file(filename)
    mapping_payload = entry.get("mapping_payload")
    if mapping_payload is None:
        raise HTTPException(status_code=400, detail="Сначала выполните нормализацию через POST /api/normalize")

    mapping_bytes = _build_mapping_excel(mapping_payload)
    base = Path(filename).stem
    return StreamingResponse(
        io.BytesIO(mapping_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{base}__mapping.xlsx"'},
    )


@app.post("/api/anomalies")
async def anomalies(req: AnomaliesRequest):
    entry = _require_file(req.filename)
    sheets_data: dict[str, pd.DataFrame] = entry["sheets_data"]

    result: dict[str, list[dict]] = {}
    for sh in req.sheets:
        df = sheets_data.get(sh)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Лист '{sh}' не найден")
        groups = scan_anomalies(df, sample_size=req.sample_size)
        result[sh] = [
            {
                "key": g.key,
                "title": g.title,
                "severity": g.severity,
                "description": g.description,
                "count": g.count,
                "examples": [
                    {"row": e.row, "column": e.column, "value": str(e.value) if e.value is not None else None}
                    for e in g.examples
                ],
            }
            for g in groups
        ]

    return result


# ---------------------------------------------------------------------------
# Document endpoints
# ---------------------------------------------------------------------------

@app.post("/api/docs/upload")
async def docs_upload(file: UploadFile = File(...)):
    ext = Path(file.filename or "").suffix.lower().lstrip(".")
    supported = ("txt", "docx", "md", "rtf")
    if ext not in supported:
        raise HTTPException(status_code=400, detail=f"Поддерживаются: {', '.join(supported)}")

    data = await file.read()
    filename = file.filename

    try:
        doc = extract_document(filename, data)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

    _doc_cache[filename] = {"doc": doc, "matches": None}

    return {"filename": filename, "text": doc.full_text}


@app.post("/api/docs/scan")
async def docs_scan(req: DocScanRequest):
    entry = _doc_cache.get(req.filename)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"Документ '{req.filename}' не найден в кэше")

    doc = entry["doc"]
    matches = scan_text_document(doc)
    entry["matches"] = matches

    grouped = group_by_type(matches)
    result: dict[str, list[dict]] = {}
    for data_type, values in grouped.items():
        # Deduplicate while preserving order for display
        seen: set[str] = set()
        unique: list[str] = []
        for v in values:
            if v not in seen:
                seen.add(v)
                unique.append(v)

        # Find spans for each unique value
        spans: dict[str, list[dict]] = {}
        for m in matches:
            if m.data_type == data_type:
                spans.setdefault(m.value, []).append({"chunk_idx": m.chunk_idx, "start": m.start, "end": m.end})

        result[data_type] = [
            {"text": v, "start": spans[v][0]["start"], "end": spans[v][0]["end"]}
            for v in unique
        ]

    return {"groups": result}


@app.post("/api/docs/normalize")
async def docs_normalize(req: DocNormalizeRequest):
    entry = _doc_cache.get(req.filename)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"Документ '{req.filename}' не найден в кэше")

    doc = entry["doc"]
    matches = entry.get("matches")
    if matches is None:
        raise HTTPException(status_code=400, detail="Сначала выполните сканирование через POST /api/docs/scan")

    # Build mapping {data_type: {original: canonical}}
    mapping: dict[str, dict[str, str]] = {}
    for data_type, selections in req.selections.items():
        type_map: dict[str, str] = {}
        for sel in selections:
            if not sel.apply:
                continue
            canonical = (sel.canonical or "").strip()
            if not canonical:
                continue
            for v in sel.variants:
                type_map[v] = canonical
        if type_map:
            mapping[data_type] = type_map

    replaced_chunks, _ = apply_replacements(doc, matches, mapping)
    result_bytes, result_ext = rebuild_document(doc, replaced_chunks)

    base = Path(req.filename).stem
    out_filename = f"{base}__normalized.{result_ext}"
    media_type = (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        if result_ext == "docx"
        else "text/plain"
    )

    return StreamingResponse(
        io.BytesIO(result_bytes),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{out_filename}"'},
    )
